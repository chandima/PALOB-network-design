"""Generate AP Placement Supplementary Information Excel file from an OpenIntent JSON export.

Usage:
    python -m src.generate_excel <path_to_openintent_json> <output_excel_path>

The script will:
  * Read the OpenIntent JSON.
  * Group access points by their 'floorplan_name'.
  * Create one worksheet per floor, matching the header format of the COOR example.
  * Populate rows with sensible defaults for antenna, mounting, etc.
  * Save the resulting workbook to the desired location.

Author: ChatGPT
"""
import json
import sys
import re
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font

# Header row copied from the COOR example (trailing empty columns removed)
HEADER = [
    "Placement Number",
    "AP Model",
    "Antenna Type",
    "Antenna Vendor",
    "Antenna Model",
    "Mount Type",
    "Mounting Bracket",
    "Mounting Adapter",
    "AP in Enclosure?",
    "Enclosure Model",
    "Antenna in Enclosure?",
    "Direction *",
    "Downtilt *",
    "Notes",
]

# Simple conversion from floorplan names to sheet names
FLOOR_RE = re.compile(r"^(\d+)(?:st|nd|rd|th)?\s+Floor$", re.I)


def normalise_sheet_name(name: str) -> str:
    match = FLOOR_RE.match(name.strip())
    if match:
        return f"Floor {match.group(1)}"
    # Fall‑back: title‑case and truncate to 31 chars (Excel limit)
    return name.title()[:31]


def main() -> None:
    if len(sys.argv) != 3:
        print("Usage: python -m src.generate_excel <input.json> <output.xlsx>")
        sys.exit(1)

    json_path = Path(sys.argv[1]).expanduser()
    out_path = Path(sys.argv[2]).expanduser()

    with json_path.open() as f:
        data = json.load(f)

    aps = data.get("accesspoints", [])
    if not aps:
        print("No access points found in JSON – aborting.")
        sys.exit(1)

    # Group APs by floor
    by_floor = defaultdict(list)
    for ap in aps:
        floor = ap.get("floorplan_name", "Unknown Floor")
        by_floor[floor].append(ap)

    wb = Workbook()
    # Remove default sheet created by openpyxl
    wb.remove(wb.active)

    placement_counter = 1
    for floor_name, aps_on_floor in by_floor.items():
        sheet_name = normalise_sheet_name(floor_name)
        ws = wb.create_sheet(sheet_name)
        ws.append(HEADER)

        # Make header row bold
        bold_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold_font

        for idx, ap in enumerate(aps_on_floor, start=1):
            placement_number = f"AP-{placement_counter:03d}"
            placement_counter += 1

            model_raw = ap.get("model", "")
            model = model_raw.upper().replace("C", "C-") if model_raw.lower(
            ).startswith("c") and "-" not in model_raw else model_raw.upper()

            # Coordinates to include in notes (optional)
            coord = ap.get("coordinate_xyz") or {}
            coord_note = f"x={coord.get('x'):.1f}, y={coord.get('y'):.1f}" if coord else ""

            row = [
                placement_number,
                model,
                "Internal Omni",
                "N/A",
                "N/A",
                "Hard Ceiling",
                "<insert mounting bracket model>",
                "<insert mounting adapter model>",
                "No",
                "N/A",
                "N/A",
                "N/A",
                "N/A",
                coord_note,
            ]
            ws.append(row)

        # Freeze header row
        ws.freeze_panes = "A2"

    wb.save(out_path)
    print(f"Excel file written to {out_path}")


if __name__ == "__main__":
    main()
